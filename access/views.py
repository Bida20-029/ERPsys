from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .forms import leaveForm, paymentForm, pettyForm, ApprovalForm
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.forms import UserCreationForm
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from .models import Payment, Petty, Employee, LeaveRequest
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.contrib import messages
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.conf import settings


import csv
from .forms import LeaveRequestForm
from .models import LeaveRequest
from django.contrib.auth.decorators import login_required


@login_required
def create_leave_request(request):
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.user = request.user
            leave_request.save()
            return redirect('confirm_leave_request', leave_request_id=leave_request.id)
    else:
        form = LeaveRequestForm()
    return render(request, 'employees/leave_form.html', {'form': form})


@login_required
def leave_requests(request):
    leave_requests = LeaveRequest.objects.filter(user=request.user)
    return render(request, 'employees/leave_requests.html', {'leave_requests': leave_requests})

def pay_request(request):
    pay_requests = Payment.objects.filter(Voucher_ID=request.Voucher_ID)
    return render(request, 'payments/single_payment.html', {'pay_requests': pay_requests})

@login_required
def submit_leave_request(request):
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST)
        if form.is_valid():
            leave_request = form.save(commit=False)
            leave_request.user = request.user
            leave_request.confirmation_status = 'P'  # Set confirmation status to 'Pending'
            leave_request.save()

            # Send email notification to admin
            subject = 'Leave Request Confirmation'
            message = render_to_string('employees/leave_request_email.html', {'leave_request': leave_request})
            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [settings.EMAIL_HOST_USER])

            messages.success(request, 'Leave request submitted successfully.')
            return redirect('confirm_leave_request')
    else:
        form = LeaveRequestForm()
    return render(request, 'employees/submit_leave_request.html', {'form': form})

from django.contrib.admin.views.decorators import staff_member_required

@staff_member_required
def confirm_leave_request(request, leave_request_id):
    leave_request = get_object_or_404(LeaveRequest, id=leave_request_id)
    
    if request.method == 'POST':
        leave_request.confirmation_status = 'A'  # Set confirmation status to 'Approved'
        leave_request.save()

        # Deduct confirmed days from employee's available leave days
        leave_request.available_leave_days -= leave_request.requested_days
        leave_request.save()
        
        messages.success(request, 'Leave request confirmed successfully.')
        return redirect('employees')

    return render(request, 'employees/confirm_leave_request.html', {'leave_request': leave_request})


leave_requestslist = [
    {
        'leave_request_id':'1',
        'user':'Gaone',
        'requested_days': '2',
        'available_leave_days':'28',
        'start_date':'06/06/2023',
        'end_date':'07/06/2023',
        'reason':'Sick-leave',
        'confirmation_status': 'A'
    }
]

employeelist = [
    {
        'Emp_ID':'1',
        'Name':'Gaone',
        'Surname':'Dimakatso',
        'leave_START':'10-April-2023',
        'leave_END':'10-April-2023',
        'Approved': '1',
    }
]

paymentlist = [
    {
        'Voucher_ID':'1',
        'Prepare_By':'Gaone Dimakatso',
        'Funds_use_breakdown':'Laptops',
        'Amount_in_words':'19,000',
    }
]

pettylist = [
    {
        'Petty_ID':'',
        'Date':'',
        'Receipt':'',
        'Cash_Used_For':'',
        'Total_Price':'',
        'Petty_Cash_Taken':'',
        'Petty_Cash_Used':'',
        'Balance':'',
        'Balance_c/d':'',
        'Filled_In_By':'',
    }
]

def employees(request):
    context = {'employees':employeelist}
    return render(request, 'employees/employees.html', context)

def leave_req(request):
    context = {'employees':employeelist}
    return render(request, 'employees/employees.html', context)

def employee(request, pk):
    moduleObj = None
    for i in employeelist:
        if i['Emp_ID'] == pk:
            moduleObj = i
            
    return render(request, 'employees/leave_requests.html', {'employee':moduleObj})

def payments(request):
    context = {'payments':paymentlist}
    return render(request, 'payments/payments.html', context)

def payment(request, pk):
    moduleObj = None
    for i in paymentlist:
        if i['Voucher_ID'] == pk:
            moduleObj = i
            
    return render(request, 'payments/single-payment.html', {'payment':moduleObj})

def pettys(request):
    context = {'pettys':pettylist}
    return render(request, 'payments/payments.html', context)

def petty(request, pk):
    moduleObj = None
    for i in pettylist:
        if i['Petty_ID'] == pk:
            moduleObj = i
            
    return render(request, 'Pettys/single-petty.html', {'petty':moduleObj})

def createEmployee(request):
     form = leaveForm()
     if request.method == 'POST':
         form = leaveForm(request.POST, request.FILES)
         if form.is_valid():
          form.save()
          return redirect('employees')
      
     context = {'form': form}
     return render(request, "employees/leave_form.html", context) 

def createPayment(request):
    form = paymentForm()
    if request.method == 'POST':
        form = paymentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            
            # subject = 'Payment Voucher Confirmation'
            #message = render_to_string('payments/payment_form.html', {'form': form})
            #message = f"A new payment voucher has been drawn up, please log in and approve as soon as possible."
            #send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [settings.EMAIL_HOST_USER])

            #messages.success(request, 'Leave request submitted successfully.')
            return redirect('employees')
        
    context = {'form': form}
    return render(request, 'payments/payment_form.html', context)

def createPetty(request):
    form = pettyForm()
    if request.method == 'POST':
        form = pettyForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('employees')
        
    context = {'form': form}
    return render(request, 'Pettys/Petty_form.html', context)
 

def pdf_leave(request):
    path = "static/pdfs/Leaveapplication.pdf"

    # Open the file in binary mode
    with open(path, 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline;filename=Leaveapplication.pdf'
        return response
    
def pdf_payment(request):
    path = "static/pdfs/PaymentVoucher.pdf"

    # Open the file in binary mode
    with open(path, 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline;filename=PaymentVoucher.pdf'
        return response
    
    
def get_petty(request, pk):
 moduleObj = petty.object.get(pk)
 return render(request, 'petty/single_petty.html', ({'petty':moduleObj}))
# Create your views here.
    

def export_to_csv(request):
    mydata = Payment.objects.all()
    response = HttpResponse('text/csv; charset=utf-8')
    response ['Content-Disposition'] = 'attachment; filename=payment_export.csv'
    writer = csv.writer(response)
    writer.writerow(['Voucher_ID', 'Prepared_By', 'Funds_Use_Breakdown', 'Amount_In_Words'])
    mydata_fields = mydata.values_list('Voucher_ID', 'Prepared_By', 'Funds_use_breakdown', 'Amount_in_words')
    for payment in mydata_fields:
        writer.writerow(payment)
        return response
    
def export_to_petty(request):
    mydata = Petty.objects.all()
    response = HttpResponse('text/csv; charset=utf-8')
    response ['Content-Disposition'] = 'attachment; filename=payment_export.csv'
    writer = csv.writer(response)
    writer.writerow(['Petty_ID', 'Date', 'Cash_Used_For', 'Total_Price', 'Petty_Cash_Taken', ' Petty_Cash_Used', 'Balance', 'Filled_In_By'])
    mydata_fields = mydata.values_list('Petty_ID', 'Date', 'Cash_Used_For', 'Total_Price', 'Petty_Cash_Taken', ' Petty_Cash_Used', 'Balance', 'Filled_In_By')
    for mydata in mydata_fields:
        writer.writerow(Petty)
        return response
    
def petty_to_excel(request):
    mydata = Petty.objects.all()
    # Load existing Excel file
    wb = openpyxl.load_workbook('petty.xlsx')
    # Select the first worksheet
    ws = wb.active
    # Determine the next empty row
    row = ws.max_row + 1
    # Write the header row if this is the first time writing to the file
    if row == 2:
        ws.cell(row=1, column=1).value = 'Petty_ID'
        ws.cell(row=1, column=2).value = 'Date'
        ws.cell(row=1, column=3).value = 'Cash_Used_For'
        ws.cell(row=1, column=4).value = 'Total_Price'
        ws.cell(row=1, column=5).value = 'Petty_Cash_Taken'
        ws.cell(row=1, column=6).value = 'Petty_Cash_Used'
        ws.cell(row=1, column=7).value = 'Balance'
        ws.cell(row=1, column=8).value = 'Filled_In_By'
    # Write data to the worksheet
    for petty in mydata:
        ws.cell(row=row, column=1).value = petty.Petty_ID
        ws.cell(row=row, column=2).value = petty.Date
        ws.cell(row=row, column=3).value = petty.Cash_Used_For
        ws.cell(row=row, column=4).value = petty.Total_Price
        ws.cell(row=row, column=5).value = petty.Petty_Cash_Taken
        ws.cell(row=row, column=6).value = petty.Petty_Cash_Used
        ws.cell(row=row, column=7).value = petty.Balance
        ws.cell(row=row, column=8).value = petty.Filled_In_By
        row += 1
    # Save changes to the Excel file
    wb.save('petty.xlsx')
    # Return response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="petty.xlsx"'
    return response

def payment_to_excel(request):
    mydata = Payment.objects.all()
    # Load existing Excel file
    wb = openpyxl.load_workbook('payment.xlsx')
    # Select the first worksheet
    ws = wb.active
    # Determine the next empty row
    row = ws.max_row + 1
    # Write the header row if this is the first time writing to the file
    if row == 2:
        ws.cell(row=1, column=1).value = 'VoucherID'
        ws.cell(row=1, column=2).value = 'PaidTo'
        ws.cell(row=1, column=3).value = 'PreparedBy'
        ws.cell(row=1, column=4).value = 'FundsUseBreakdown'
        ws.cell(row=1, column=5).value = 'AmountInWords'
        ws.cell(row=1, column=6).value = 'Amount'
        ws.cell(row=1, column=7).value = 'Date'
    # Write data to the worksheet
    for payment in mydata:
        ws.cell(row=row, column=1).value = payment.Voucher_ID
        ws.cell(row=row, column=2).value = payment.Paid_To
        ws.cell(row=row, column=3).value = payment.Prepared_By
        ws.cell(row=row, column=4).value = payment.Funds_use_breakdown
        ws.cell(row=row, column=5).value = payment.Amount_in_words
        ws.cell(row=row, column=6).value = payment.Amount
        ws.cell(row=row, column=7).value = payment.Date
        row += 1
    # Save changes to the Excel file
    wb.save('payment.xlsx')
    # Return response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="payment.xlsx"'
    return response



