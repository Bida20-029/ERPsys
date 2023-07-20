from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.http import FileResponse
from .forms import loanForm, memberForm, cellForm, CreateUserForm
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.forms import UserCreationForm
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, get_object_or_404
from .models import Voucher, Loan, Member
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
import csv
import pyodbc
from django.core.management.base import BaseCommand
import pandas as pd
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

def registerPage(request):
    form = CreateUserForm()
    
    if request.method == 'POST':
        form = CreateUserForm(request.POST)
        if form.is_valid:
            form.save()
            user = form.cleaned_data.get('username')
            messages.success(request, 'Account for' + user + 'successfully created')
            return redirect('login')
             
    context = {'form':form}
    return render(request, 'signup.html', context)

from django.contrib.auth import authenticate, login

def loginPage(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request)  # Corrected line
            return redirect('main')
        else:
            messages.info(request, 'Username or password is incorrect')

    context = {}
    return render(request, 'Login.html', context)

def logoutUser(request):
    logout(request)
    return redirect('login')

def voucher_to_sql(request):
    mydata = Member.objects.all()

    # Connect to the SQL Server database
    cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=<server_name>;DATABASE=<database_name>;UID=<username>;PWD=<password>')
    cursor = cnxn.cursor()

    # Iterate through each member and insert their data into the SQL database
    for member in mydata:
        cursor.execute("INSERT INTO <table_name> (ID_Num, Name, Surname, LandBoard, Contact, Principal, Residential_Address) VALUES (?, ?, ?, ?, ?, ?, ?)",
            member.ID_Num, member.Name, member.Surname, member.LandBoard, member.Contact, member.Principal, member.Residential_Address)

    # Commit the changes to the database and close the cursor and database connection
    cursor.commit()
    cursor.close()
    cnxn.close()

    # Return a success message to the user
    return HttpResponse("Data exported to SQL database successfully!")

voucherlist = [
    {
        'IDNum':'1',
        'Name':'Anita',
        'Surname': 'Len',
        'Contact': '72345273',
        'Voucher': 'P1,800',
        'Cell_Amount':'1,800',
        'Offer': '75',
        'Voucher_Start':'March-2023',
        'Voucher_End': 'Feb-2025',
        'Form':'Insert form'
    }
]

myapplist = [
    {
        'app':'1'
    }
]

employeelist = [
    {
        'Emp_ID':'1',
        'Name':'Gaone',
        'Surname':'Dimakatso',
        'leavedays':'42',
        'Form':'Insert form'
    }
]

loanlist = [
    {
        'Loan_ID':'12345678',
        'Name':'Gaone',
        'Surname':'Dimakatso',
        'Duration':'2years',
        'LandBoard':'Tlokweng',
        'Principal':'25',
        'START':'December-2022',
        'END':'November-2024',
        'Form':'Insert form'
        }
]

memberlist = [
    {
        'ID_no':'169728918',
        'Name':'Gaone',
        'Surname':'Dimakatso',
        'LandBoard':'Tlokweng',
        'Contact':'72870098',
        'Residential_Address':'',
        'Form':'Insert form'
    }
]

paymentlist = [
    {
        'Voucher_ID':'1',
        'Name':'Gaone',
        'Surname':'Dimakatso',
        'Reason':'Laptops',
        'Amount':'19,000',
        'Form':'Insert from'
    }
]

celllist = [
    {
        'ID_Num':'169728918',
        'Name':'Gaone',
        'Surname':'Dimakatso',
        'LandBoard':'Tlokweng',
        'Form':'Insert from'
    }
]

def home(request):
    return render(request, 'index.html')

def about(request):
    return render(request, 'About.html')

@login_required
def main(request):
    return render(request, 'test.html')

def myapp(request,pk):
    moduleObj = None
    for i in myapplist:
        if i['myapp'] == pk:
            moduleObj = i
            
    return render(request, 'myapps/single-app.html', {'myapp':moduleObj})

def loans(request):
    context = {'loans':loanlist}
    return render(request, 'loans/loans.html', context)

def members(request):
    context = {'members':memberlist}
    return render(request, 'members/members.html', context)

def member(request, pk):
    moduleObj = None
    for i in memberlist:
        if i['ID_no'] == pk:
            moduleObj = i
            
    return render(request, 'members/single-member.html', {'member':moduleObj})

def vouchers(request):
    context = {'cells':celllist}
    return render(request, 'vouchers/vouchers.html', context)

def voucher(request,pk):
    moduleObj = None
    for i in myapplist:
        if i['ID_Num'] == pk:
            moduleObj = i
            
    return render(request, 'vouchersys/single-voucher.html', {'voucher':moduleObj})

def createLoan(request):
    form = loanForm()
    if request.method == 'POST':
        form = loanForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('loans')
        
    context = {'form': form}
    return render(request, "loan_form.html", context)

def createVoucher(request):
    form = cellForm()
    if request.method == 'POST':
        form = cellForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('vouchers')
        
    context = {'form': form}
    return render(request, "vouchersys/voucher_form.html", context) 

def createMember(request):
    form = memberForm()
    if request.method == 'POST':
        form = memberForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('members')
        
    context = {'form':form}
    return render(request, "members/member_form.html", context) 

def VouchersPage(request):
    return render(request, 'vouchersys/VouchersPage.html')

def LoansPage(request):
    return render(request, "loans/loanspage.html")

def MembersPage(request):
    return render(request, "members/memberspage.html")

def MainPage(request):
    return render(request, "Main.html")

def login(request):
    form = AuthenticationForm(request)
    return render(request, 'login.html', {'form': form})

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})

def pdf_view(request):
    # Path to the PDF file
    path = "static/pdfs/Union_Hybrid_Voice_Application_Form_20220328 (2).pdf"

    # Open the file in binary mode
    with open(path, 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline;filename=Union_Hybrid_Voice_Application_Form_20220328 (2).pdf'
        return response
    
def pdf_loan(request):
    path = "static/pdfs/LOANFORM.pdf"

    # Open the file in binary mode
    with open(path, 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline;filename=LOANFORM.pdf'
        return response
    
def pdf_member(request):
    path = "static/pdfs/Membershipform.pdf"

    # Open the file in binary mode
    with open(path, 'rb') as pdf:
        response = HttpResponse(pdf.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline;filename=Mmembershipform.pdf'
        return response

def get_voucher(request, pk):
    voucher = voucher.object.get(vouchers, ID_Num=pk)
    return render(request, 'vouchersys/single_voucher.html', {'voucher': voucher})

def voucher_to_csv(request):
    mydata = Voucher.objects.all()
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="mydata.csv"'
    writer = csv.writer(response)
    writer.writerow(['ID_Num', 'Name', 'Surname', 'Contact', 'Voucher', 'Cell_Amount', 'Offer', 'Voucher_Start', 'Voucher_End', 'Form'])
    mydata_fields = mydata.values_list('ID_Num', 'Name', 'Surname', 'Contact', 'Voucher', 'Cell_Amount', 'Offer', 'Voucher_Start', 'Voucher_End', 'Form')
    for row in mydata_fields:
        writer.writerow(row)
    return response

def loan_to_csv(request):
    mydata = Loan.objects.all()
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="mydata.csv"'
    writer = csv.writer(response)
    writer.writerow(['Loan_ID', 'Name', 'Surname', 'Duration', 'LandBoard', 'Principal', 'Start', 'End'])
    mydata_fields = mydata.values_list('Loan_ID', 'Name', 'Surname', 'Duration', 'LandBoard', 'Principal', 'Start', 'End')
    for row in mydata_fields:
        writer.writerow(row)
    return response

def member_to_csv(request):
    mydata = Member.objects.all()
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="mydata.xlsx"'
    writer = csv.writer(response)
    writer.writerow(['ID_Num', 'Name', 'Surname', 'LandBoard',  'Contact', 'Principal', 'Residential_Address'])
    mydata_fields = mydata.values_list('ID_Num', 'Name', 'Surname', 'LandBoard',  'Contact', 'Principal', 'Residential_Address')
    for row in mydata_fields:
        writer.writerow(row)
    return response

def voucher_to_excel(request):
    mydata = Voucher.objects.all()
    # Load existing Excel file
    wb = openpyxl.load_workbook('voucher.xlsx')
    # Select the first worksheet
    ws = wb.active
    # Determine the next empty row
    row = ws.max_row + 1
    # Write the header row if this is the first time writing to the file
    if row == 2:
        ws.cell(row=1, column=2).value = 'Name'
        ws.cell(row=1, column=3).value = 'Surname'
        ws.cell(row=1, column=4).value = 'ID_Number'
        ws.cell(row=1, column=5).value = 'Contact'
        ws.cell(row=1, column=6).value = 'Voucher'
        ws.cell(row=1, column=7).value = 'Cell_Amount'
        ws.cell(row=1, column=8).value = 'Offer'
        ws.cell(row=1, column=9).value = 'Start_'
        ws.cell(row=1, column=10).value = 'End_'
    # Write data to the worksheet
    for voucher in mydata:
        ws.cell(row=row, column=2).value = voucher.Name
        ws.cell(row=row, column=3).value = voucher.Surname
        ws.cell(row=row, column=4).value = voucher.ID_Num
        ws.cell(row=row, column=5).value = voucher.Contact
        ws.cell(row=row, column=6).value = voucher.Voucher
        ws.cell(row=row, column=7).value = voucher.Cell_Amount
        ws.cell(row=row, column=8).value = voucher.Offer
        ws.cell(row=row, column=9).value = voucher.Voucher_Start
        ws.cell(row=row, column=10).value = voucher.Voucher_End
        row += 1
    # Save changes to the Excel file
    wb.save('voucher.xlsx')
    # Return response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="voucher.xlsx"'
    return response

def loan_to_excel(request):
    mydata = Loan.objects.all()
    # Load existing Excel file
    wb = openpyxl.load_workbook('loan.xlsx')
    # Select the first worksheet
    ws = wb.active
    # Determine the next empty row
    row = ws.max_row + 1
    # Write the header row if this is the first time writing to the file
    if row == 2:
        ws.cell(row=1, column=1).value = 'Loan_ID'
        ws.cell(row=1, column=2).value = 'Name'
        ws.cell(row=1, column=3).value = 'Surname'
        ws.cell(row=1, column=4).value = 'Duration'
        ws.cell(row=1, column=5).value = 'Land_board'
        ws.cell(row=1, column=6).value = 'Principal'
        ws.cell(row=1, column=7).value = 'Start_'
        ws.cell(row=1, column=8).value = 'End_'
    # Write data to the worksheet
    for loan in mydata:
        ws.cell(row=row, column=1).value = loan.Loan_ID
        ws.cell(row=row, column=2).value = loan.Name
        ws.cell(row=row, column=3).value = loan.Surname
        ws.cell(row=row, column=4).value = loan.Duration
        ws.cell(row=row, column=5).value = loan.LandBoard
        ws.cell(row=row, column=6).value = loan.Principal
        ws.cell(row=row, column=7).value = loan.Start
        ws.cell(row=row, column=8).value = loan.End
        row += 1
    # Save changes to the Excel file
    wb.save('loan.xlsx')
    # Return response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="loan.xlsx"'
    return response

class Command(BaseCommand):
    help = 'Exports data from Django models to a flat file'

    def handle(self, *args, **options):
        queryset = Voucher.objects.all()  # Replace 'YourModel' with your actual model name

        # Convert queryset to a pandas DataFrame
        df = pd.DataFrame(list(queryset.values()))

        # Export DataFrame to a flat file
        file_path = 'vouchers.csv'  # Specify the path where you want to save the file
        df.to_csv(file_path, index=False)  # You can use different file formats and options here
        self.stdout.write(self.style.SUCCESS(f'Data exported successfully to {file_path}'))
        
        
        
